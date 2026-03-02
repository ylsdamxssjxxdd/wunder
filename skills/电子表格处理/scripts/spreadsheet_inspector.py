#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Spreadsheet inspector for unknown/possibly broken files.

Purpose:
- Quickly detect whether a spreadsheet file is readable.
- Surface likely issues (format mismatch, corruption, potential encryption).
- Output structured metadata for LLM workflows (JSON), reducing model guesswork.

Supported:
- .xlsx/.xlsm/.xltx/.xltm
- .csv/.tsv
- .xls (best-effort, depends on local xlrd availability)
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zipfile import BadZipFile

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - best effort for runtime env
    load_workbook = None


ZIP_BASED_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
TEXT_EXTS = {".csv", ".tsv"}
OLE_HEADER = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
ZIP_HEADER = b"PK\x03\x04"


@dataclass
class Issue:
    severity: str  # info/warning/error
    code: str
    message: str


def read_file_header(path: Path, n: int = 8) -> bytes:
    with path.open("rb") as f:
        return f.read(n)


def detect_encoding(path: Path, candidates: List[str]) -> Tuple[Optional[str], Optional[str]]:
    for enc in candidates:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                f.read(4096)
            return enc, None
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            return None, f"read_error({enc}): {exc}"
    return None, "no_suitable_encoding"


def detect_csv_delimiter(path: Path, encoding: str) -> str:
    try:
        with path.open("r", encoding=encoding, newline="") as f:
            sample = f.read(8192)
        sniffed = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        return sniffed.delimiter
    except Exception:
        # fallback by extension
        return "\t" if path.suffix.lower() == ".tsv" else ","


def safe_to_numeric(series: pd.Series) -> pd.Series:
    s = series.astype(str)
    s = s.str.replace(",", "", regex=False).str.replace("%", "", regex=False).str.strip()
    return pd.to_numeric(s, errors="coerce")


def infer_column_role(col_name: str, series: pd.Series) -> Dict[str, Any]:
    non_null = series.dropna()
    nn = int(non_null.shape[0])
    total = int(series.shape[0])
    null_ratio = (1 - nn / total) if total else 1.0

    sample = str(non_null.iloc[0]) if nn else ""
    nunique = int(non_null.nunique(dropna=True)) if nn else 0
    unique_ratio = (nunique / nn) if nn else 0.0

    id_name_hit = bool(re.search(r"(id|编号|员工号|工号|订单号|流水号)$", col_name, re.IGNORECASE))
    time_name_hit = bool(re.search(r"(date|time|日期|时间|月份|周期)", col_name, re.IGNORECASE))
    metric_name_hit = bool(re.search(r"(amount|qty|count|rate|比例|金额|数量|时长|天数|次数|占比|率)", col_name, re.IGNORECASE))

    numeric_ratio = 0.0
    datetime_ratio = 0.0
    if nn:
        num = safe_to_numeric(non_null)
        numeric_ratio = float(num.notna().mean())

        # Avoid misclassifying plain numeric columns as datetime.
        text = non_null.astype(str)
        has_date_tokens_ratio = float(text.str.contains(r"[-/年月日Tt:]", regex=True).mean())
        should_try_datetime = bool(time_name_hit or (numeric_ratio < 0.7 and has_date_tokens_ratio >= 0.3))
        if should_try_datetime:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=UserWarning)
                dt = pd.to_datetime(non_null, errors="coerce")
            datetime_ratio = float(dt.notna().mean())

    role = "category_candidate"
    confidence = 0.55

    if (id_name_hit and unique_ratio > 0.7) or (unique_ratio > 0.95 and nn > 20):
        role = "id_candidate"
        confidence = 0.88
    elif datetime_ratio >= 0.8 or (time_name_hit and datetime_ratio >= 0.5):
        role = "time_candidate"
        confidence = 0.9 if datetime_ratio >= 0.8 else 0.75
    elif numeric_ratio >= 0.8 or (metric_name_hit and numeric_ratio >= 0.5):
        role = "metric_candidate"
        confidence = 0.9 if numeric_ratio >= 0.8 else 0.72

    return {
        "name": col_name,
        "dtype": str(series.dtype),
        "null_ratio": round(null_ratio, 4),
        "sample": sample,
        "non_null_count": nn,
        "nunique_non_null": nunique,
        "unique_ratio": round(unique_ratio, 4),
        "numeric_ratio": round(numeric_ratio, 4),
        "datetime_ratio": round(datetime_ratio, 4),
        "role_guess": role,
        "role_confidence": round(confidence, 4),
    }


def profile_dataframe(df: pd.DataFrame) -> Dict[str, Any]:
    columns = [infer_column_role(str(c), df[c]) for c in df.columns]
    role_map = {
        "id_cols": [c["name"] for c in columns if c["role_guess"] == "id_candidate"],
        "time_cols": [c["name"] for c in columns if c["role_guess"] == "time_candidate"],
        "metric_cols": [c["name"] for c in columns if c["role_guess"] == "metric_candidate"],
        "category_cols": [c["name"] for c in columns if c["role_guess"] == "category_candidate"],
    }
    return {"columns": columns, "role_candidates": role_map}


def quick_health_checks(df: pd.DataFrame, sheet_name: str) -> List[Issue]:
    issues: List[Issue] = []
    if df.shape[0] == 0:
        issues.append(Issue("warning", "EMPTY_SHEET", f"Sheet '{sheet_name}' has 0 rows in sampled data."))
    if df.shape[1] == 0:
        issues.append(Issue("error", "NO_COLUMNS", f"Sheet '{sheet_name}' has 0 columns."))
        return issues

    col_names = [str(c) for c in df.columns]
    dup = pd.Series(col_names).duplicated()
    if dup.any():
        repeated = pd.Series(col_names)[dup].tolist()
        issues.append(
            Issue("warning", "DUPLICATE_COLUMNS", f"Sheet '{sheet_name}' has duplicated columns: {repeated[:10]}")
        )

    unnamed = [c for c in col_names if c.lower().startswith("unnamed")]
    if unnamed and len(unnamed) / max(len(col_names), 1) > 0.3:
        issues.append(
            Issue(
                "warning",
                "TOO_MANY_UNNAMED_COLUMNS",
                f"Sheet '{sheet_name}' has many unnamed columns ({len(unnamed)}/{len(col_names)}).",
            )
        )

    high_null_cols = df.columns[df.isna().mean() > 0.95].tolist()
    if high_null_cols:
        issues.append(
            Issue(
                "info",
                "HIGH_NULL_COLUMNS",
                f"Sheet '{sheet_name}' has high-null columns: {list(map(str, high_null_cols[:10]))}",
            )
        )
    return issues


def _header_candidate_score(columns: List[str]) -> Tuple[float, int, float, int]:
    if not columns:
        return (1.0, 10**9, 1.0, -1)
    col_s = pd.Series(columns)
    unnamed_ratio = float(col_s.str.lower().str.startswith("unnamed").mean())
    dup_count = int(col_s.duplicated().sum())
    numeric_like_ratio = float(col_s.str.fullmatch(r"\d+(\.\d+)?").fillna(False).mean())
    unique_count = int(col_s.nunique(dropna=False))
    # Lower is better for first 3 fields, higher is better for unique_count.
    return (unnamed_ratio, dup_count, numeric_like_ratio, -unique_count)


def detect_best_header_row(path: Path, sheet_name: str, sample_rows: int, max_scan_rows: int = 5) -> Tuple[int, pd.DataFrame]:
    best_header = 0
    best_df: Optional[pd.DataFrame] = None
    best_score: Optional[Tuple[float, int, float, int]] = None

    for header_row in range(0, max_scan_rows + 1):
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                nrows=sample_rows,
                dtype=object,
                engine="openpyxl",
                header=header_row,
            )
        except Exception:
            continue
        cols = [str(c) for c in df.columns]
        score = _header_candidate_score(cols)
        if best_score is None or score < best_score:
            best_score = score
            best_header = header_row
            best_df = df

    if best_df is None:
        # Let caller handle by raising later with default read behavior.
        best_df = pd.read_excel(path, sheet_name=sheet_name, nrows=sample_rows, dtype=object, engine="openpyxl")
        best_header = 0
    return best_header, best_df


def classify_open_error(exc: Exception) -> Tuple[str, str]:
    msg = str(exc)
    low = msg.lower()
    if isinstance(exc, BadZipFile) or "badzipfile" in low:
        return "error", "File is not a valid zip-based Excel package (.xlsx/.xlsm), possibly corrupted or wrong suffix."
    if "password" in low or "encrypted" in low:
        return "error", "File appears encrypted/password-protected."
    if "not a zip file" in low:
        return "error", "File extension suggests xlsx, but content is not zip. Possibly encrypted/legacy/corrupt."
    if "unsupported format" in low or "invalid file" in low:
        return "error", "Unsupported or invalid Excel format."
    return "error", f"Failed to open file: {msg}"


def inspect_excel(path: Path, sample_rows: int, selected_sheets: Optional[List[str]]) -> Dict[str, Any]:
    issues: List[Issue] = []
    workbook_meta: Dict[str, Any] = {"sheet_names": [], "sheets": []}

    if load_workbook is None:
        issues.append(Issue("error", "OPENPYXL_MISSING", "openpyxl is unavailable in current environment."))
        return {"ok": False, "issues": [asdict(i) for i in issues], "meta": workbook_meta}

    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        sev, msg = classify_open_error(exc)
        issues.append(Issue(sev, "EXCEL_OPEN_FAILED", msg))
        return {"ok": False, "issues": [asdict(i) for i in issues], "meta": workbook_meta}

    sheet_names = wb.sheetnames
    workbook_meta["sheet_names"] = sheet_names

    target_sheets = selected_sheets if selected_sheets else sheet_names
    missing_sheets = [s for s in target_sheets if s not in sheet_names]
    if missing_sheets:
        issues.append(
            Issue("warning", "MISSING_SHEETS", f"Requested sheets not found and skipped: {missing_sheets}")
        )
    target_sheets = [s for s in target_sheets if s in sheet_names]

    for sheet in target_sheets:
        ws = wb[sheet]
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        try:
            header_row, df = detect_best_header_row(path, sheet_name=sheet, sample_rows=sample_rows)
        except Exception as exc:
            sev, msg = classify_open_error(exc)
            issues.append(Issue(sev, "SHEET_READ_FAILED", f"Sheet '{sheet}' read failed: {msg}"))
            continue

        profile = profile_dataframe(df)
        sheet_issues = quick_health_checks(df, sheet)
        if header_row > 0:
            sheet_issues.append(
                Issue(
                    "warning",
                    "HEADER_ROW_SHIFTED",
                    f"Sheet '{sheet}' likely uses header row {header_row + 1}, not row 1.",
                )
            )
        issues.extend(sheet_issues)

        workbook_meta["sheets"].append(
            {
                "name": sheet,
                "header_row_detected": int(header_row),
                "max_row_estimate": max_row,
                "max_col_estimate": max_col,
                "sampled_rows": int(df.shape[0]),
                "sampled_cols": int(df.shape[1]),
                "profile": profile,
                "issues": [asdict(i) for i in sheet_issues],
            }
        )

    return {"ok": not any(i.severity == "error" for i in issues), "issues": [asdict(i) for i in issues], "meta": workbook_meta}


def inspect_text_table(path: Path, sample_rows: int) -> Dict[str, Any]:
    issues: List[Issue] = []
    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk", "utf-16", "latin1"]
    encoding, err = detect_encoding(path, encodings)
    if not encoding:
        issues.append(Issue("error", "ENCODING_DETECT_FAILED", f"Encoding detection failed: {err}"))
        return {"ok": False, "issues": [asdict(i) for i in issues], "meta": {"sheet_names": [], "sheets": []}}

    sep = detect_csv_delimiter(path, encoding)
    try:
        df = pd.read_csv(path, sep=sep, nrows=sample_rows, dtype=object, encoding=encoding, engine="python")
    except Exception as exc:
        issues.append(Issue("error", "TEXT_READ_FAILED", f"Failed to read table text file: {exc}"))
        return {"ok": False, "issues": [asdict(i) for i in issues], "meta": {"sheet_names": [], "sheets": []}}

    profile = profile_dataframe(df)
    sheet_issues = quick_health_checks(df, "table")
    issues.extend(sheet_issues)

    meta = {
        "sheet_names": ["table"],
        "encoding": encoding,
        "delimiter": sep,
        "sheets": [
            {
                "name": "table",
                "sampled_rows": int(df.shape[0]),
                "sampled_cols": int(df.shape[1]),
                "profile": profile,
                "issues": [asdict(i) for i in sheet_issues],
            }
        ],
    }
    return {"ok": not any(i.severity == "error" for i in issues), "issues": [asdict(i) for i in issues], "meta": meta}


def build_top_level_suggestions(result: Dict[str, Any]) -> List[str]:
    issues = result.get("issues", [])
    issue_codes = {i["code"] for i in issues}
    suggestions: List[str] = []

    if "EXCEL_OPEN_FAILED" in issue_codes:
        suggestions.append("If file is password-protected, ask user for decrypted copy or password workflow.")
    if "ENCODING_DETECT_FAILED" in issue_codes:
        suggestions.append("Try exporting source file to UTF-8 CSV and rerun.")
    if "TOO_MANY_UNNAMED_COLUMNS" in issue_codes:
        suggestions.append("Likely merged headers; confirm actual header row before analysis.")

    # Generic recommendations for LLMs
    suggestions.append("Use role_candidates to build role_map before aggregation/charting.")
    suggestions.append("Prefer read-only analysis mode unless user explicitly asks for write-back.")
    return suggestions


def summarize_stdout(report: Dict[str, Any]) -> None:
    print("=== Spreadsheet Inspector ===")
    print(f"path: {report['path']}")
    print(f"exists: {report['exists']}, readable: {report['readable']}, status: {report['status']}")
    print(f"file_type: {report.get('file_type')}, size_bytes: {report.get('size_bytes')}")
    print(f"header_signature: {report.get('header_signature')}")

    if report.get("issues"):
        print("issues:")
        for i in report["issues"]:
            print(f"  - [{i['severity']}] {i['code']}: {i['message']}")
    else:
        print("issues: none")

    meta = report.get("meta", {})
    sheets = meta.get("sheets", [])
    if sheets:
        print(f"sheets ({len(sheets)}):")
        for s in sheets:
            name = s["name"]
            rows = s.get("sampled_rows")
            cols = s.get("sampled_cols")
            rc = s["profile"]["role_candidates"]
            print(
                f"  - {name}: sampled_rows={rows}, sampled_cols={cols}, "
                f"time={rc['time_cols'][:3]}, metric={rc['metric_cols'][:5]}, category={rc['category_cols'][:5]}"
            )
    else:
        print("sheets: none")

    if report.get("suggestions"):
        print("suggestions:")
        for s in report["suggestions"]:
            print(f"  - {s}")


def inspect(path: Path, sample_rows: int, selected_sheets: Optional[List[str]]) -> Dict[str, Any]:
    report: Dict[str, Any] = {
        "path": str(path),
        "exists": path.exists(),
        "readable": False,
        "status": "error",
        "file_type": path.suffix.lower(),
        "size_bytes": path.stat().st_size if path.exists() else 0,
        "header_signature": "unknown",
        "issues": [],
        "meta": {"sheet_names": [], "sheets": []},
        "suggestions": [],
    }
    if not path.exists():
        report["issues"] = [asdict(Issue("error", "FILE_NOT_FOUND", "Input file does not exist."))]
        report["suggestions"] = ["Check absolute/relative path and rerun."]
        return report

    try:
        header = read_file_header(path)
    except Exception as exc:
        report["issues"] = [asdict(Issue("error", "FILE_READ_FAILED", str(exc)))]
        report["suggestions"] = ["Check file permissions."]
        return report

    suffix = path.suffix.lower()
    if header.startswith(ZIP_HEADER):
        report["header_signature"] = "zip"
    elif header.startswith(OLE_HEADER):
        report["header_signature"] = "ole"
    else:
        report["header_signature"] = "other"

    if suffix in ZIP_BASED_EXTS and report["header_signature"] == "ole":
        report["issues"].append(
            asdict(
                Issue(
                    "warning",
                    "SIGNATURE_EXTENSION_MISMATCH",
                    "File extension is xlsx-like but binary signature is OLE. It may be encrypted, legacy, or mislabeled.",
                )
            )
        )
    elif suffix in ZIP_BASED_EXTS and report["header_signature"] != "zip":
        report["issues"].append(
            asdict(
                Issue(
                    "warning",
                    "SIGNATURE_EXTENSION_MISMATCH",
                    "File extension is xlsx-like but signature is not zip. It may be invalid/corrupt.",
                )
            )
        )

    if suffix in ZIP_BASED_EXTS or suffix == ".xls":
        excel_result = inspect_excel(path, sample_rows=sample_rows, selected_sheets=selected_sheets)
        report["meta"] = excel_result["meta"]
        report["issues"].extend(excel_result["issues"])
        report["readable"] = bool(excel_result["ok"])
    elif suffix in TEXT_EXTS:
        text_result = inspect_text_table(path, sample_rows=sample_rows)
        report["meta"] = text_result["meta"]
        report["issues"].extend(text_result["issues"])
        report["readable"] = bool(text_result["ok"])
    else:
        report["issues"].append(
            asdict(
                Issue(
                    "error",
                    "UNSUPPORTED_SUFFIX",
                    f"Unsupported file suffix '{suffix}'. Supported: {sorted(list(ZIP_BASED_EXTS | TEXT_EXTS | {'.xls'}))}",
                )
            )
        )
        report["readable"] = False

    has_error = any(i["severity"] == "error" for i in report["issues"])
    report["status"] = "error" if has_error else ("ok" if report["readable"] else "warning")
    report["suggestions"] = build_top_level_suggestions(report)
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect spreadsheet health and emit metadata JSON.")
    parser.add_argument("path", help="Input spreadsheet path (.xlsx/.xlsm/.xls/.csv/.tsv).")
    parser.add_argument("--sample-rows", type=int, default=2000, help="Rows to sample per sheet.")
    parser.add_argument(
        "--sheet",
        action="append",
        default=None,
        help="Specific sheet(s) to inspect for Excel files. Repeatable.",
    )
    parser.add_argument(
        "--output-json",
        default=None,
        help="Output JSON path. If omitted, only stdout summary is printed.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress stdout summary. Useful when piping JSON output only.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    path = Path(args.path)
    report = inspect(path, sample_rows=max(1, args.sample_rows), selected_sheets=args.sheet)

    if args.output_json:
        out_path = Path(args.output_json)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    if not args.quiet:
        summarize_stdout(report)
        if args.output_json:
            print(f"json_written: {args.output_json}")

    return 0 if report["status"] != "error" else 2


if __name__ == "__main__":
    sys.exit(main())
